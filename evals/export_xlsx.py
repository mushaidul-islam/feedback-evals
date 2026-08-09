#!/usr/bin/env python3
"""Collapse a run directory into one workbook, one worksheet per model.

    python3 evals/export_xlsx.py evals/results/run_20260809T101500Z

Writes results.xlsx into the run directory:

    summary            score_summary.csv, if score_results.py has been run
    <model label>      every response from that model
    disagreements      wrong or invalid responses across all models
    rewrites           every category-2 rewrite produced

The CSVs remain the source of truth; this is for reading and sorting by hand.

Requires openpyxl:  pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

# Excel sheet names cannot exceed 31 characters or contain these.
FORBIDDEN = set(r"[]:*?/\\")
MAX_WIDTH = 60


def sheet_name(raw: str, used: set[str]) -> str:
    cleaned = "".join("-" if character in FORBIDDEN else character for character in raw)[:31]
    candidate, suffix = cleaned or "sheet", 2
    while candidate.lower() in used:
        candidate = f"{cleaned[:28]}_{suffix}"
        suffix += 1
    used.add(candidate.lower())
    return candidate


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def add_sheet(workbook: Workbook, title: str, header: list[str], rows: list[list[Any]],
              used: set[str]) -> None:
    sheet = workbook.create_sheet(sheet_name(title, used))
    sheet.append(header)
    for row in rows:
        sheet.append(row)

    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top")
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = sheet.dimensions

    for index, column in enumerate(header, start=1):
        longest = max([len(str(column))] + [len(str(row[index - 1])) for row in rows[:400]
                                            if index - 1 < len(row)])
        sheet.column_dimensions[get_column_letter(index)].width = min(MAX_WIDTH, max(10, longest + 2))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("run_directory", type=Path)
    parser.add_argument("--output", type=Path, default=None,
                        help="Workbook path (default: <run directory>/results.xlsx)")
    args = parser.parse_args()

    run_directory = args.run_directory
    if not run_directory.is_dir():
        print(f"Not a directory: {run_directory}", file=sys.stderr)
        return 1

    combined = run_directory / "combined.csv"
    if not combined.is_file():
        print(f"No combined.csv in {run_directory} — run the eval first.", file=sys.stderr)
        return 1

    header, rows = read_csv(combined)
    label_index = header.index("model_label")

    workbook = Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()

    summary = run_directory / "score_summary.csv"
    if summary.is_file():
        add_sheet(workbook, "summary", *read_csv(summary), used=used)
    else:
        print("No score_summary.csv yet — run score_results.py for a summary sheet.")

    run_config = run_directory / "run.json"
    if run_config.is_file():
        config = json.loads(run_config.read_text(encoding="utf-8"))
        flat = [[key, json.dumps(value, ensure_ascii=False)[:32000] if not isinstance(value, str) else value]
                for key, value in config.items()]
        add_sheet(workbook, "run_config", ["key", "value"], flat, used=used)

    by_label: dict[str, list[list[str]]] = defaultdict(list)
    for row in rows:
        by_label[row[label_index]].append(row)
    for label in sorted(by_label):
        add_sheet(workbook, label, header, by_label[label], used=used)

    for extra in ("disagreements", "rewrites_for_review"):
        path = run_directory / f"{extra}.csv"
        if path.is_file():
            add_sheet(workbook, extra, *read_csv(path), used=used)

    output = args.output or run_directory / "results.xlsx"
    workbook.save(output)
    print(f"Wrote {output} — {len(by_label)} model sheet(s), {len(rows)} responses.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
